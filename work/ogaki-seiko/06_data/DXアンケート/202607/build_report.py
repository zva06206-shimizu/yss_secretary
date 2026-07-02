# -*- coding: utf-8 -*-
"""共有用の整形xlsxを生成する。
- スコア表：Responsesを整形（不要列削除・列名整理・回答を点数化・■除去）
- 集計サマリ：カテゴリ別／部署別／拠点別／全社スコア＋評価基準
- 採点基準：各設問の選択肢→点数（重み）と採点ルール
採点は analyze.py と同一（5=成熟, 該当なし除外）。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, BarChart, Reference
import analyze as A  # SCALE, CATEGORIES, score, load_responses, COL_* を再利用

SHORT = {  # グラフ用の短いカテゴリ名
 '1.ペーパーレス化':'ペーパーレス','2.データ一元化':'データ一元化','3.集計・転記の自動化':'集計自動化',
 '4.システム連携':'システム連携','5.情報・ナレッジ共有':'情報共有','6.承認・意思決定スピード':'承認スピード',
}

SRC = 'DX推進状況アンケート ご協力のお願い.xlsx'
OUT = 'DXアンケート集計_共有用_中間20260702.xlsx'

HEAD_FILL = PatternFill('solid', fgColor='1F4E78')
HEAD_FONT = Font(color='FFFFFF', bold=True)
SUB_FILL  = PatternFill('solid', fgColor='D9E1F2')
BOLD = Font(bold=True)
CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

Q_HEADERS = {  # 設問の表示名（点数列の見出し）
 'Q1':'Q1 日常業務の帳票の紙運用割合','Q2':'Q2 稟議・申請・承認の紙＋押印割合',
 'Q3':'Q3 紙書類の探索・保管に週何時間','Q4':'Q4 データがどこにあるか分からない頻度',
 'Q5':'Q5 同一データの重複管理割合','Q6':'Q6 拠点間リアルタイム共有の割合',
 'Q7':'Q7 月次・日報集計の手作業割合','Q8':'Q8 転記・入力ミスの月間発生回数',
 'Q9':'Q9 定例レポート作成時間(月合計)','Q10':'Q10 システム間データ連携状況',
 'Q11':'Q11 複数システムへの二重入力頻度','Q12':'Q12 手順書・標準作業書の電子化',
 'Q13':'Q13 属人化の程度','Q14':'Q14 稟議申請〜承認完了の日数',
 'Q15':'Q15 経営数字が把握できるまでの時間',
}
CAT_STATE = lambda c: "要改善" if c<2.0 else "改善余地大" if c<3.0 else "概ね良好" if c<4.0 else "定着・活用"

def style_header(ws, row, ncol):
    for j in range(1, ncol+1):
        c = ws.cell(row=row, column=j); c.fill=HEAD_FILL; c.font=HEAD_FONT; c.alignment=CEN; c.border=BORDER

def build():
    data = A.load_responses()
    N = len(data)
    wb = openpyxl.Workbook()

    # ---- シート1：スコア表 ----
    ws = wb.active; ws.title = 'スコア表(点数化)'
    headers = ['No.','回答ID','拠点名','部署名','お名前'] + [Q_HEADERS[q] for q in A.SCALE] \
              + ['他部門でも同じ入力を感じるか','どのような業務でそう感じるか','DX推進・全般のご意見']
    ws.append(headers); style_header(ws, 1, len(headers))
    for c in data:
        row = [c.get(1,''), c.get(2,''), c.get(A.COL_SITE,''), c.get(A.COL_DEPT,''), c.get(9,'')]
        for q in A.SCALE:
            sc = A.score(q, c.get(A.COL_Q1+int(q[1:])-1))
            row.append(sc if sc is not None else '')   # 該当なし/わからない は空欄(除外)
        row += [c.get(25,''), c.get(26,''), c.get(27,'')]
        ws.append(row)
    widths = [5,10,22,12,14] + [8]*15 + [22,30,40]
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes='F2'
    ws.cell(row=ws.max_row+2, column=1, value='※ 点数は 5=成熟(良い)〜1=未成熟(悪い)。「該当なし/わからない」は空欄＝集計から除外。')

    # ---- シート2：集計サマリ ----
    s = wb.create_sheet('集計サマリ')
    r = 1
    s.cell(r,1,'大垣精工 DX推進状況アンケート 集計サマリ（中間）').font=Font(bold=True,size=14); r+=1
    s.cell(r,1,f'回答数 N={N}（2026-07-02時点・回答期限 7/10 の途中経過）'); r+=2

    # 設問別平均
    qmean={}; qn={}
    for q in A.SCALE:
        sc=[A.score(q, c.get(A.COL_Q1+int(q[1:])-1)) for c in data]
        valid=[x for x in sc if x is not None]; qmean[q]=A.mean(sc); qn[q]=len(valid)
    # カテゴリ平均・全社スコア
    cat={}
    for name,qs in A.CATEGORIES:
        ms=[qmean[q] for q in qs if qmean[q] is not None]
        cat[name]=round(sum(ms)/len(ms),2) if ms else None
    total=round(sum(((cat[n]-1)/4*16.67) for n,_ in A.CATEGORIES if cat[n] is not None),1)
    band = "紙・属人化中心" if total<=25 else "電子化途上" if total<=50 else "業務改善定着" if total<=75 else "データ活用・自動化段階"

    s.cell(r,1,'■ 全社DX成熟度スコア').font=BOLD; r+=1
    s.cell(r,1,'スコア'); s.cell(r,2,total); s.cell(r,3,'／100'); s.cell(r,4,f'状態：{band}')
    for j in range(1,5): s.cell(r,j).font=Font(bold=True,size=12)
    r+=2

    # 評価基準（追加要望）
    s.cell(r,1,'■ 評価基準').font=BOLD; r+=1
    s.cell(r,1,'全社スコア帯'); s.cell(r,2,'状態'); style_header(s,r,2); r+=1
    for lo,hi,st in [(0,25,'紙・属人化中心'),(26,50,'電子化途上'),(51,75,'業務改善定着'),(76,100,'データ活用・自動化段階')]:
        s.cell(r,1,f'{lo}〜{hi}'); s.cell(r,2,st)
        for j in (1,2): s.cell(r,j).border=BORDER
        r+=1
    r+=1
    s.cell(r,1,'カテゴリ平均の状態'); s.cell(r,2,'目安'); style_header(s,r,2); r+=1
    for rng,st in [('1.0〜1.9','要改善'),('2.0〜2.9','改善余地大'),('3.0〜3.9','概ね良好'),('4.0〜5.0','定着・活用段階')]:
        s.cell(r,1,rng); s.cell(r,2,st)
        for j in (1,2): s.cell(r,j).border=BORDER
        r+=1
    r+=1

    # カテゴリ別平均
    s.cell(r,1,'■ カテゴリ別平均（5点満点・該当なし除外・低い順）').font=BOLD; r+=1
    s.cell(r,1,'カテゴリ'); s.cell(r,2,'平均'); s.cell(r,3,'状態'); s.cell(r,4,'対象設問'); style_header(s,r,4); r+=1
    for name in sorted(cat, key=lambda k:(cat[k] if cat[k] is not None else 99)):
        qs=dict(A.CATEGORIES)[name]; c=cat[name]
        s.cell(r,1,name); s.cell(r,2,c); s.cell(r,3,CAT_STATE(c) if c is not None else '-'); s.cell(r,4,'・'.join(qs))
        for j in range(1,5): s.cell(r,j).border=BORDER
        r+=1
    r+=1

    # 部署別・拠点別
    def group(colidx):
        g={}
        for c in data:
            k=c.get(colidx,'(未記入)') or '(未記入)'
            sc=[A.score(q,c.get(A.COL_Q1+int(q[1:])-1)) for q in A.SCALE]; sc=[x for x in sc if x is not None]
            g.setdefault(k,[0,0,0]); g[k][0]+=sum(sc); g[k][1]+=len(sc); g[k][2]+=1
        return g
    for title,colidx in [('■ 部署別 総合平均',A.COL_DEPT),('■ 拠点別 総合平均',A.COL_SITE)]:
        s.cell(r,1,title+'（全設問・該当なし除外）').font=BOLD; r+=1
        s.cell(r,1,'区分'); s.cell(r,2,'平均'); s.cell(r,3,'回答者数'); s.cell(r,4,'有効回答数'); style_header(s,r,4); r+=1
        g=group(colidx)
        for k,v in sorted(g.items(), key=lambda kv:-(kv[1][0]/kv[1][1] if kv[1][1] else 0)):
            avg=round(v[0]/v[1],2) if v[1] else None
            s.cell(r,1,k); s.cell(r,2,avg); s.cell(r,3,v[2]); s.cell(r,4,v[1])
            for j in range(1,5): s.cell(r,j).border=BORDER
            r+=1
        r+=1

    # 設問別平均
    s.cell(r,1,'■ 設問別平均（5点満点・該当なし除外）').font=BOLD; r+=1
    s.cell(r,1,'設問'); s.cell(r,2,'平均'); s.cell(r,3,'有効n'); style_header(s,r,3); r+=1
    for q in A.SCALE:
        s.cell(r,1,Q_HEADERS[q]); s.cell(r,2,qmean[q]); s.cell(r,3,f'{qn[q]}/{N}')
        for j in range(1,4): s.cell(r,j).border=BORDER
        r+=1
    for col,w in {'A':30,'B':10,'C':12,'D':26}.items(): s.column_dimensions[col].width=w
    s.cell(r+1,1,'※ Zoho内蔵の平均(Statistics)は「該当なし」を最高点として数えるため使用せず、該当なしを除外して算出。').font=Font(italic=True,size=9)

    # ---- シート3：採点基準（重み） ----
    w3 = wb.create_sheet('採点基準(重み)')
    w3.append(['設問','点数1（最も未成熟）','点数2','点数3','点数4','点数5（最も成熟）','除外']); style_header(w3,1,7)
    for q in A.SCALE:
        w3.append([Q_HEADERS[q]] + A.SCALE[q] + ['該当なし/わからない'])
    for j,w in enumerate([30,16,14,14,14,18,16],1): w3.column_dimensions[get_column_letter(j)].width=w
    for row in w3.iter_rows(min_row=1, max_row=w3.max_row, max_col=7):
        for c in row: c.border=BORDER; c.alignment=LEFT
    w3.cell(w3.max_row+2,1,'※ 5=成熟(良い)〜1=未成熟(悪い)。「該当なし/わからない」は集計から除外。').font=Font(italic=True,size=9)

    # ---- シート4：設問別平均データ ----
    w4 = wb.create_sheet('設問別平均データ')
    w4.append(['設問','平均点','有効回答数(n)','全回答数(N)']); style_header(w4,1,4)
    for q in A.SCALE:
        w4.append([Q_HEADERS[q], qmean[q], qn[q], N])
    for row in w4.iter_rows(min_row=1,max_row=w4.max_row,max_col=4):
        for c in row: c.border=BORDER; c.alignment=LEFT
    for j,w in enumerate([32,10,14,10],1): w4.column_dimensions[get_column_letter(j)].width=w

    # ---- シート5：グラフ（カテゴリ平均の積み上げ棒＋レーダー） ----
    g = wb.create_sheet('グラフ')
    cat_order = [(SHORT[n], cat[n]) for n,_ in A.CATEGORIES]
    # レーダー用データ（縦持ち）
    g.cell(1,1,'カテゴリ'); g.cell(1,2,'平均点'); style_header(g,1,2)
    for i,(nm,av) in enumerate(cat_order, start=2):
        g.cell(i,1,nm); g.cell(i,2,av)
        g.cell(i,1).border=BORDER; g.cell(i,2).border=BORDER
    g.column_dimensions['A'].width=16; g.column_dimensions['B'].width=10
    # 積み上げ用データ（横持ち：1本の棒に6カテゴリを積む）
    hdr_row=1; val_row=2; base=5  # E列(5)以降
    g.cell(hdr_row, base-1, '')  # D1相当は空
    g.cell(val_row, base-1, 'DX成熟度')
    for k,(nm,av) in enumerate(cat_order):
        g.cell(hdr_row, base+k, nm)
        g.cell(val_row, base+k, av)

    # レーダーチャート
    radar = RadarChart(); radar.type='filled'; radar.style=26
    radar.title='DX成熟度 レーダーチャート（カテゴリ平均・5点満点）'
    rdata = Reference(g, min_col=2, min_row=1, max_row=1+len(cat_order))
    rcats = Reference(g, min_col=1, min_row=2, max_row=1+len(cat_order))
    radar.add_data(rdata, titles_from_data=True); radar.set_categories(rcats)
    radar.height=10; radar.width=13
    g.add_chart(radar, 'A10')

    # 積み上げ棒グラフ
    bar = BarChart(); bar.type='col'; bar.grouping='stacked'; bar.overlap=100
    bar.title='カテゴリ別平均点 積み上げ'
    bdata = Reference(g, min_col=base, max_col=base+len(cat_order)-1, min_row=hdr_row, max_row=val_row)
    bcats = Reference(g, min_col=base-1, min_row=val_row, max_row=val_row)
    bar.add_data(bdata, titles_from_data=True, from_rows=True); bar.set_categories(bcats)
    bar.y_axis.title='平均点(積み上げ)'; bar.height=10; bar.width=13
    g.add_chart(bar, 'J10')

    wb.save(OUT)
    print('saved:', OUT, 'N=', N)

if __name__=='__main__':
    build()
